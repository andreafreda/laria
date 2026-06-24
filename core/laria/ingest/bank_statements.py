"""Parse BancoPosta and Postepay statement exports into importable movements.

Pure, testable logic: given the rows of an export (xlsx or csv) as a list of cell
lists, it recognizes the format, normalizes each movement to
{date (ISO), amount (signed), description, category} and computes a dedup hash.
No network, no database here; feed the result to ``finance.import_transactions``.

Two profiles:
  - BancoPosta: separate "Addebiti"/"Accrediti" columns, amount = credits - debits
  - Postepay:   a single signed "Importo" column (negative = outflow)

The keyword list matches Italian bank descriptions, so the needles stay Italian;
the categories they map to are LARIA's English category names.
"""
from __future__ import annotations

import datetime
import hashlib
import re

# Italian description keyword (lower) to LARIA category. First match wins.
# "transfer" marks internal BancoPosta -> Postepay top-ups so they are not
# mistaken for real spending.
_CATEGORY_KEYWORDS: list[tuple[tuple[str, ...], str]] = [
    (("ricarica postepay", "ricarica carta"), "transfer"),
    (("prelievo", "prelevamento", "prel. ", "atm"), "cash withdrawal"),
    (("f24", "pagopa", "pago pa", "tributi", "imposta", "tasse", "bollo", "canone rai"), "taxes"),
    (("amazon", "zalando", "ebay", "aliexpress", "shein"), "shopping"),
    (("glovo", "just eat", "justeat", "deliveroo", "uber eats"), "dining"),
    (("tim ", "vodafone", "windtre", "wind tre", "iliad", "fastweb"), "subscriptions"),
    (("netflix", "spotify", "disney", "prime video", "dazn", "youtube"), "subscriptions"),
    (("eni", "q8", "ip ", "esso", "tamoil", "benzina", "carburante", "distributore"), "fuel"),
    (("esselunga", "conad", "coop", "lidl", "eurospin", "carrefour", "pam ",
      "supermercato", "md ", "penny", "decò", "deco "), "groceries"),
    (("trenitalia", "italo", "atac", "gtt", "autostrad", "telepass", "pedaggio",
      "parcheggio", "bus ", "metro"), "transport"),
    (("farmacia", "parafarmacia", "medic", "ticket sanitario", "dottor"), "health"),
    (("commission", "competenze", "spese tenuta", "imposta di bollo"), "bank fees"),
    (("stipendio", "emolumenti", "cedolino", "salario"), "salary"),
    (("enel", "eni gas", "a2a", "hera", "acea", "iren", "bolletta", "luce", "gas"), "utilities"),
]


def categorize(description: str, amount: float) -> str:
    """Guess a category from the description.

    Unrecognized income becomes "other income", unrecognized spending "misc".
    """
    text = (description or "").lower()
    for needles, category in _CATEGORY_KEYWORDS:
        if any(needle in text for needle in needles):
            return category
    return "other income" if amount > 0 else "misc"


def parse_date(value) -> str | None:
    """Normalize a date to ISO YYYY-MM-DD, or None if it can't be parsed.

    Accepts dd/mm/yyyy, dd-mm-yyyy, yyyy-mm-dd, and date/datetime objects (which
    openpyxl may hand back).
    """
    if value is None:
        return None
    if hasattr(value, "year") and hasattr(value, "month") and hasattr(value, "day"):
        return f"{value.year:04d}-{value.month:02d}-{value.day:02d}"
    text = str(value).strip().split(" ")[0]  # drop any time part
    if not text:
        return None
    iso = re.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})$", text)
    if iso:
        year, month, day = iso.groups()
        return _iso_or_none(int(year), int(month), int(day))
    dmy = re.match(r"^(\d{1,2})[/-](\d{1,2})[/-](\d{4})$", text)
    if dmy:
        day, month, year = dmy.groups()
        return _iso_or_none(int(year), int(month), int(day))
    return None


def _iso_or_none(year: int, month: int, day: int) -> str | None:
    """ISO string only if the date really exists (rejects 32/02 and friends)."""
    try:
        return datetime.date(year, month, day).isoformat()
    except ValueError:
        return None


def _to_amount(value) -> float | None:
    """Parse an amount cell (Italian '1.234,56' or '-12,50') to float, None if blank."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("€", "").replace(" ", "")
    if not text:
        return None
    if "," in text:  # Italian format: dot thousands, comma decimals
        text = text.replace(".", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def _norm(value) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).lower()


def _find_header(rows: list[list], needles: list[str]) -> int | None:
    """Index of the first row that contains all the needles (matched on lower cells)."""
    for index, row in enumerate(rows):
        cells = [_norm(cell) for cell in row]
        joined = " | ".join(cells)
        if all(any(n in c for c in cells) or n in joined for n in needles):
            return index
    return None


def _column(header: list, needle: str) -> int | None:
    for index, cell in enumerate(header):
        if needle in _norm(cell):
            return index
    return None


def detect_format(rows: list[list]) -> str | None:
    """Return 'bancoposta', 'postepay', or None based on the headers present."""
    if _find_header(rows, ["addebiti", "accrediti", "descrizione"]) is not None:
        return "bancoposta"
    if _find_header(rows, ["importo", "descrizione"]) is not None:
        return "postepay"
    return None


def parse(rows: list[list]) -> dict:
    """Parse raw rows into movements ready for import.

    Returns {format, account, movements, skipped} (plus {error} if the format is
    not recognized). Each movement is {date, amount, description, category, hash}.
    Identical real rows in one file get distinct hashes so the second is not
    swallowed by dedup at import time.
    """
    fmt = detect_format(rows)
    if fmt is None:
        return {"format": None, "account": None, "movements": [], "skipped": 0,
                "error": "Unrecognized format (no BancoPosta/Postepay headers)."}

    header_index, columns, account = _locate_columns(rows, fmt)
    movements: list[dict] = []
    skipped = 0
    seen_keys: dict[str, int] = {}
    for row in rows[header_index + 1:]:
        if _is_blank(row):
            continue
        movement = _parse_row(row, fmt, columns, account, seen_keys)
        if movement is None:
            skipped += 1
            continue
        movements.append(movement)
    return {"format": fmt, "account": account, "movements": movements, "skipped": skipped}


def _locate_columns(rows: list[list], fmt: str) -> tuple[int, dict, str]:
    """Find the header row and the column indexes each format needs."""
    if fmt == "bancoposta":
        header_index = _find_header(rows, ["addebiti", "accrediti", "descrizione"])
        header = rows[header_index]
        columns = {
            "date": _date_column(header),
            "debit": _column(header, "addebiti"),
            "credit": _column(header, "accrediti"),
            "description": _column(header, "descrizione"),
        }
        return header_index, columns, "bancoposta"

    header_index = _find_header(rows, ["importo", "descrizione"])
    header = rows[header_index]
    columns = {
        "date": _date_column(header),
        "amount": _column(header, "importo"),
        "description": _column(header, "descrizione"),
    }
    return header_index, columns, "postepay"


def _date_column(header: list) -> int | None:
    """Prefer the booking date column, falling back to a plain "data" column.

    Uses an explicit None check because column 0 is a valid index that a simple
    ``or`` would wrongly skip.
    """
    booking = _column(header, "data contabile")
    return booking if booking is not None else _column(header, "data")


def _is_blank(row: list) -> bool:
    return not row or all((c is None or str(c).strip() == "") for c in row)


def _cell(row: list, index: int | None):
    """Cell at ``index`` if present, else None (guards short or missing columns)."""
    if index is None or index >= len(row):
        return None
    return row[index]


def _parse_row(row: list, fmt: str, columns: dict, account: str,
               seen_keys: dict[str, int]) -> dict | None:
    """Turn one data row into a movement, or None if it can't be used."""
    date = parse_date(_cell(row, columns["date"]))
    if date is None:
        return None
    description = str(_cell(row, columns["description"]) or "").strip()

    amount = _row_amount(row, fmt, columns)
    if amount is None:
        return None
    amount = round(amount, 2)

    return {
        "date": date,
        "amount": amount,
        "description": description,
        "category": categorize(description, amount),
        "hash": _dedup_hash(account, date, amount, description, seen_keys),
    }


def _row_amount(row: list, fmt: str, columns: dict) -> float | None:
    """The signed amount for a row, per the format's columns."""
    if fmt == "postepay":
        return _to_amount(_cell(row, columns["amount"]))
    debit = _to_amount(_cell(row, columns["debit"]))
    credit = _to_amount(_cell(row, columns["credit"]))
    if debit is None and credit is None:
        return None
    return (credit or 0.0) - abs(debit or 0.0)


def _dedup_hash(account: str, date: str, amount: float, description: str,
                seen_keys: dict[str, int]) -> str:
    """Stable hash for dedup; later duplicates in the same file get a suffix.

    The first occurrence uses the bare historical hash (so re-importing a file
    stays idempotent); repeats append a counter so genuine identical rows survive.
    """
    key = f"{account}|{date}|{amount:.2f}|{_norm(description)}"
    seen = seen_keys.get(key, 0)
    seen_keys[key] = seen + 1
    if seen == 0:
        return row_hash(account, date, amount, description)
    return hashlib.sha1(f"{key}|{seen}".encode("utf-8")).hexdigest()


def row_hash(account: str, date: str, amount: float, description: str) -> str:
    """Stable hash for a movement so the same row re-imported is not duplicated."""
    key = f"{account}|{date}|{amount:.2f}|{_norm(description)}"
    return hashlib.sha1(key.encode("utf-8")).hexdigest()


# --- file readers (xlsx via openpyxl, csv via the stdlib) ------------------

def rows_from_csv(content: bytes) -> list[list]:
    """Read CSV bytes into rows, autodetecting comma or semicolon delimiter."""
    import csv
    import io
    text = content.decode("utf-8-sig", errors="replace")
    sample = text[:2048]
    delimiter = ";" if sample.count(";") >= sample.count(",") else ","
    return [row for row in csv.reader(io.StringIO(text), delimiter=delimiter)]


def rows_from_xlsx(content: bytes) -> list[list]:
    """Read xlsx bytes into rows. Requires the optional ``openpyxl`` dependency."""
    import io
    from openpyxl import load_workbook
    # read_only=False: some Poste exports declare a range not starting at A1,
    # which breaks row iteration in read-only mode.
    workbook = load_workbook(io.BytesIO(content), read_only=False, data_only=True)
    worksheet = workbook.active
    rows = [list(r) for r in worksheet.iter_rows(values_only=True)]
    workbook.close()
    return rows


def rows_from_file(content: bytes, filename: str) -> list[list]:
    """Pick the reader by file extension (csv/txt vs xlsx)."""
    name = (filename or "").lower()
    if name.endswith(".csv") or name.endswith(".txt"):
        return rows_from_csv(content)
    return rows_from_xlsx(content)
